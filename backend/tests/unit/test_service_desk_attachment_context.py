"""Focused tests for bounded attachment context used by Service Desk AI."""

from __future__ import annotations

import base64
import io
import json
import zipfile

import pytest
from openpyxl import Workbook

from aexy.services import gmail_sync_service as gmail_sync
from aexy.services.gmail_sync_service import GmailSyncService


def test_csv_preview_contains_header_and_three_sample_rows():
    raw = b"policy_no,member_name\nP-1,Asha\nP-2,Ravi\nP-3,Neha\nP-4,Ignored\n"

    preview = GmailSyncService._service_desk_preview("borrowers.csv", "text/csv", raw)

    assert json.loads(preview or "[]") == [
        ["policy_no", "member_name"],
        ["P-1", "Asha"],
        ["P-2", "Ravi"],
        ["P-3", "Neha"],
    ]


def test_xlsx_preview_contains_header_and_three_sample_rows():
    workbook = Workbook()
    sheet = workbook.active
    for row in (
        ("invoice", "amount"),
        ("I-1", 100),
        ("I-2", 200),
        ("I-3", 300),
        ("I-4", 400),
    ):
        sheet.append(row)
    content = io.BytesIO()
    workbook.save(content)

    preview = GmailSyncService._service_desk_preview(
        "invoice-register.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content.getvalue(),
    )

    assert json.loads(preview or "[]") == [
        ["invoice", "amount"],
        ["I-1", "100"],
        ["I-2", "200"],
        ["I-3", "300"],
    ]


def test_text_preview_is_limited_to_six_hundred_characters():
    preview = GmailSyncService._service_desk_preview(
        "note.txt",
        "text/plain",
        b"x" * 700,
    )

    assert preview == "x" * 600


def test_pdf_is_metadata_only_at_this_context_tier():
    assert not GmailSyncService._supports_service_desk_preview(
        "invoice.pdf",
        "application/pdf",
    )


@pytest.mark.asyncio
async def test_raw_byte_ceiling_rejects_attachment_before_download(monkeypatch):
    service = GmailSyncService(None)
    download_called = False

    async def fake_request(*args, **kwargs):
        nonlocal download_called
        download_called = True
        return {"data": ""}

    monkeypatch.setattr(gmail_sync, "_SERVICE_DESK_ATTACHMENT_RAW_BYTE_LIMIT", 8)
    monkeypatch.setattr(service, "_make_gmail_request", fake_request)

    with pytest.raises(ValueError, match="raw-byte limit"):
        await service._gmail_attachment_bytes(
            object(),
            "gmail-message-1",
            {"attachmentId": "attachment-1", "size": 9},
        )

    assert download_called is False


@pytest.mark.asyncio
async def test_missing_size_rejects_external_attachment_before_download(monkeypatch):
    service = GmailSyncService(None)
    download_called = False

    async def fake_request(*args, **kwargs):
        nonlocal download_called
        download_called = True
        return {"data": ""}

    monkeypatch.setattr(service, "_make_gmail_request", fake_request)

    with pytest.raises(ValueError, match="unavailable before download"):
        await service._gmail_attachment_bytes(
            object(),
            "gmail-message-without-size",
            {"attachmentId": "attachment-without-size"},
        )

    assert download_called is False


@pytest.mark.asyncio
async def test_encoded_ceiling_rejects_attachment_before_decoding(monkeypatch):
    service = GmailSyncService(None)
    decode_called = False

    def fake_decode(*args, **kwargs):
        nonlocal decode_called
        decode_called = True
        return b""

    monkeypatch.setattr(gmail_sync, "_SERVICE_DESK_ATTACHMENT_RAW_BYTE_LIMIT", 8)
    monkeypatch.setattr(gmail_sync.base64, "urlsafe_b64decode", fake_decode)

    with pytest.raises(ValueError, match="raw-byte limit"):
        await service._gmail_attachment_bytes(
            object(),
            "gmail-message-2",
            {"data": "A" * 16},
        )

    assert decode_called is False


def test_compressed_xlsx_expansion_is_rejected_before_parsing(monkeypatch):
    content = io.BytesIO()
    with zipfile.ZipFile(content, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"x" * 64)

    parse_called = False

    def fake_load_workbook(*args, **kwargs):
        nonlocal parse_called
        parse_called = True
        raise AssertionError("expanded workbook should not reach openpyxl")

    monkeypatch.setattr(gmail_sync, "_SERVICE_DESK_XLSX_EXPANDED_BYTE_LIMIT", 32)
    monkeypatch.setattr(gmail_sync, "load_workbook", fake_load_workbook)

    preview = GmailSyncService._service_desk_preview(
        "oversized.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content.getvalue(),
    )

    assert preview is None
    assert parse_called is False


@pytest.mark.asyncio
async def test_preview_failure_preserves_attachment_metadata(monkeypatch):
    service = GmailSyncService(None)

    async def fail_preview(*args, **kwargs):
        raise ValueError("preview unavailable")

    monkeypatch.setattr(service, "_gmail_attachment_bytes", fail_preview)
    context = await service._service_desk_attachment_context(
        object(),
        "gmail-message-3",
        {
            "parts": [
                {
                    "filename": "members.csv",
                    "mimeType": "text/csv",
                    "body": {"attachmentId": "attachment-3", "size": 42},
                }
            ]
        },
    )

    assert context == [
        {
            "filename": "members.csv",
            "content_type": "text/csv",
            "size_bytes": 42,
            # Kept even when the preview fails: it is what lets a KAM forward
            # the file later, and it is an identifier rather than content.
            "attachment_id": "attachment-3",
        }
    ]


def test_attachment_parts_are_not_decoded_as_message_body():
    """A filename-bearing text part is an attachment, not the body.

    Decoding it here would pull the whole file into memory outside the bounded
    attachment path, and pass its contents to the classifier as if the
    requester had typed them. Nested parts must still be walked, so the real
    body inside the multipart wrapper is found.
    """
    service = GmailSyncService(None)

    def part(mime_type: str, raw: bytes, filename: str = "") -> dict:
        return {
            "mimeType": mime_type,
            "filename": filename,
            "body": {"data": base64.urlsafe_b64encode(raw).decode(), "size": len(raw)},
        }

    body_text, body_html = service._extract_body(
        {
            "mimeType": "multipart/mixed",
            "parts": [
                part("text/csv", b"policy_no\nP-1\n", filename="members.csv"),
                part("text/plain", b"attached note", filename="note.txt"),
                part("text/html", b"<p>attached page</p>", filename="page.html"),
                {
                    "mimeType": "multipart/alternative",
                    "parts": [
                        part("text/plain", b"Real body"),
                        part("text/html", b"<p>Real body</p>"),
                    ],
                },
            ],
        }
    )

    assert body_text == "Real body"
    assert body_html == "<p>Real body</p>"


@pytest.mark.asyncio
async def test_only_first_three_supported_attachments_receive_previews():
    service = GmailSyncService(None)
    parts = []
    for index in range(4):
        raw = f"preview-{index}".encode()
        parts.append(
            {
                "filename": f"attachment-{index}.txt",
                "mimeType": "text/plain",
                "body": {
                    "data": base64.urlsafe_b64encode(raw).decode(),
                    "size": len(raw),
                },
            }
        )

    context = await service._service_desk_attachment_context(
        object(),
        "gmail-message-4",
        {"parts": parts},
    )

    assert [item.get("preview") for item in context] == [
        "preview-0",
        "preview-1",
        "preview-2",
        None,
    ]


def test_attachment_still_base64_after_transfer_decode_is_decoded_again():
    """Some senders leave the transfer encoding on, so one decode is not enough.

    Without this the classifier reads an unintelligible blob instead of the
    claim rows, and the file forwarded to an insurer cannot be opened.
    """
    real = b"claim_ref,member_name\nCLM-1,Asha\nCLM-2,Ravi\n"
    still_encoded = base64.b64encode(real)

    assert GmailSyncService._decode_if_still_base64(
        still_encoded, filename="claims.csv", content_type="text/csv"
    ) == real


def test_legitimate_base64_text_file_is_not_silently_rewritten():
    raw = b"SGVsbG8gV29ybGQh"

    assert GmailSyncService._decode_if_still_base64(
        raw, filename="provider-token.txt", content_type="text/plain"
    ) == raw


def test_ordinary_file_is_never_decoded_a_second_time():
    for raw in (
        b"policy_no,member_name\nP-1,Asha\n",   # commas and newlines
        b"Office notes for September team.",    # spaces and a full stop
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR", # binary signature
        b"short",                               # below the minimum length
    ):
        assert GmailSyncService._decode_if_still_base64(raw) == raw


def test_double_encoded_csv_previews_as_readable_rows():
    real = b"claim_ref,member_name\nCLM-1,Asha\nCLM-2,Ravi\nCLM-3,Neha\n"
    decoded = GmailSyncService._decode_if_still_base64(
        base64.b64encode(real), filename="claims.csv", content_type="text/csv"
    )

    preview = GmailSyncService._service_desk_preview("claims.csv", "text/csv", decoded)

    assert json.loads(preview or "[]")[0] == ["claim_ref", "member_name"]


@pytest.mark.asyncio
async def test_forwarding_ceiling_is_used_after_gmail_transfer_decode(monkeypatch):
    service = GmailSyncService(None)
    raw = b"0123456789ab"

    monkeypatch.setattr(gmail_sync, "_SERVICE_DESK_ATTACHMENT_RAW_BYTE_LIMIT", 8)

    loaded = await service._gmail_attachment_bytes(
        object(),
        "gmail-message-forward",
        {"data": base64.urlsafe_b64encode(raw).decode(), "size": len(raw)},
        max_bytes=16,
    )

    assert loaded == raw

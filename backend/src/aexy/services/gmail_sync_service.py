"""Gmail Sync Service for syncing emails from Google."""

import base64
import csv
import io
import json
import logging
import re
import zipfile
from datetime import datetime, timedelta, timezone
from email.utils import parseaddr
from itertools import islice
from typing import Any

import httpx
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from aexy.core.config import get_settings
from aexy.models.google_integration import (
    EmailSyncCursor,
    GoogleIntegration,
    SyncedEmail,
    SyncedEmailRecordLink,
)
from aexy.models.crm import CRMRecord, CRMObject, CRMObjectType, CRMRecordRelation

logger = logging.getLogger(__name__)


# Payloads made of nothing but base64 characters are the ones some senders leave
# still encoded. Commas, spaces and most binary bytes fall outside this, so a
# genuine CSV, text file or document can never match.
_BASE64_ONLY_RE = re.compile(rb"[A-Za-z0-9+/]+={0,2}")

_SERVICE_DESK_ATTACHMENT_RAW_BYTE_LIMIT = 2 * 1024 * 1024
_SERVICE_DESK_XLSX_EXPANDED_BYTE_LIMIT = 16 * 1024 * 1024
_SERVICE_DESK_ATTACHMENT_METADATA_LIMIT = 10
_SERVICE_DESK_ATTACHMENT_PREVIEW_LIMIT = 3
_SERVICE_DESK_ATTACHMENT_PREVIEW_CHAR_LIMIT = 600
# Forwarding a file to an insurer is a different job from sampling it for the
# classifier, so it gets its own ceiling. Gmail itself rejects much above this,
# and holding the bytes in memory is what actually costs us.
_SERVICE_DESK_ATTACHMENT_FORWARD_BYTE_LIMIT = 10 * 1024 * 1024


# Default deal creation settings
DEFAULT_DEAL_SETTINGS = {
    "auto_create_deals": False,  # Master toggle
    "deal_creation_mode": "auto",  # "auto", "ai", "criteria"
    "skip_personal_domains": True,
    "default_deal_stage": "new",
    "default_deal_value": None,
    "criteria": {
        "subject_keywords": [],  # e.g., ["quote", "proposal", "pricing", "interested"]
        "body_keywords": [],
        "from_domains": [],  # Specific domains to create deals for
    },
}


async def auto_enrich_contact_from_email(
    db: AsyncSession,
    workspace_id: str,
    from_email: str | None,
    from_name: str | None,
    synced_email: "SyncedEmail",
) -> tuple[CRMRecord | None, CRMRecord | None]:
    """Auto-create or find contact from email sender and link the email.

    Also creates company record from email domain if applicable.
    Returns a tuple of (contact_record, company_record).
    """
    if not from_email:
        return None, None

    # Skip common personal email domains
    personal_domains = {"gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "icloud.com", "aol.com", "protonmail.com"}
    domain = from_email.split("@")[1].lower() if "@" in from_email else None

    # Skip if it's likely the user's own email
    try:
        int_result = await db.execute(
            select(GoogleIntegration).where(GoogleIntegration.workspace_id == workspace_id)
        )
        integration = int_result.scalar_one_or_none()
        if integration and integration.google_email and from_email.lower() == integration.google_email.lower():
            return None, None
    except Exception:
        pass

    # Get the Person object type
    person_obj_result = await db.execute(
        select(CRMObject).where(
            CRMObject.workspace_id == workspace_id,
            CRMObject.slug == "person",
        )
    )
    person_obj = person_obj_result.scalar_one_or_none()

    if not person_obj:
        # No person object type configured, skip enrichment
        return None, None

    # Get the Company object type
    company_obj_result = await db.execute(
        select(CRMObject).where(
            CRMObject.workspace_id == workspace_id,
            CRMObject.slug == "company",
        )
    )
    company_obj = company_obj_result.scalar_one_or_none()

    # Create company from domain if not a personal email
    company_record = None
    if company_obj and domain and domain not in personal_domains:
        # Check if company with this domain exists
        company_records_result = await db.execute(
            select(CRMRecord).where(
                CRMRecord.workspace_id == workspace_id,
                CRMRecord.object_id == company_obj.id,
            )
        )
        company_records = company_records_result.scalars().all()

        for record in company_records:
            if record.values:
                record_domain = record.values.get("domain", "").lower()
                record_website = record.values.get("website", "").lower()
                if domain == record_domain or domain in record_website:
                    company_record = record
                    break

        if not company_record:
            # Create a new company from domain
            from uuid import uuid4

            company_name = domain.split(".")[0].title()  # Simple name from domain
            company_record = CRMRecord(
                id=str(uuid4()),
                workspace_id=workspace_id,
                object_id=company_obj.id,
                display_name=company_name[:500],
                values={
                    "name": company_name,
                    "domain": domain,
                    "website": f"https://{domain}",
                },
                source="email_sync",
            )
            db.add(company_record)
            # Update object record count
            company_obj.record_count = (company_obj.record_count or 0) + 1
            await db.flush()
            logger.info(f"Auto-created company {company_name} from email domain")

    # Check if a person record already exists with this email
    person_records_result = await db.execute(
        select(CRMRecord).where(
            CRMRecord.workspace_id == workspace_id,
            CRMRecord.object_id == person_obj.id,
        )
    )
    person_records = person_records_result.scalars().all()

    # Find person with matching email
    existing_record = None
    for record in person_records:
        if record.values:
            for key, value in record.values.items():
                if isinstance(value, str) and value.lower() == from_email.lower():
                    existing_record = record
                    break
        if existing_record:
            break

    if not existing_record:
        # Create a new person record
        from uuid import uuid4

        # Parse the name
        first_name, last_name = None, None
        if from_name:
            name_parts = from_name.strip().split(" ", 1)
            first_name = name_parts[0] if name_parts else None
            last_name = name_parts[1] if len(name_parts) > 1 else None

        values = {
            "email": from_email,
        }
        if first_name:
            values["first_name"] = first_name
        if last_name:
            values["last_name"] = last_name
        if from_name:
            values["name"] = from_name

        display_name = from_name or from_email

        existing_record = CRMRecord(
            id=str(uuid4()),
            workspace_id=workspace_id,
            object_id=person_obj.id,
            display_name=display_name[:500],
            values=values,
            source="email_sync",
        )
        db.add(existing_record)
        # Update object record count
        person_obj.record_count = (person_obj.record_count or 0) + 1
        await db.flush()
        logger.info(f"Auto-created contact {display_name} from email sync")

        # Link person to company if we have one
        if company_record:
            from aexy.models.crm import CRMRecordRelation

            relation = CRMRecordRelation(
                id=str(uuid4()),
                source_record_id=existing_record.id,
                target_record_id=company_record.id,
                relation_type="works_at",
            )
            db.add(relation)

    # Link the email to this person record
    link_result = await db.execute(
        select(SyncedEmailRecordLink).where(
            SyncedEmailRecordLink.email_id == synced_email.id,
            SyncedEmailRecordLink.record_id == existing_record.id,
        )
    )
    existing_link = link_result.scalar_one_or_none()

    if not existing_link:
        from uuid import uuid4 as uuid4_link

        link = SyncedEmailRecordLink(
            id=str(uuid4_link()),
            email_id=synced_email.id,
            record_id=existing_record.id,
            link_type="from",
            confidence=1.0,
        )
        db.add(link)

    return existing_record, company_record


async def auto_create_or_update_deal_from_email(
    db: AsyncSession,
    workspace_id: str,
    synced_email: "SyncedEmail",
    contact_record: CRMRecord | None,
    company_record: CRMRecord | None,
    deal_settings: dict,
) -> CRMRecord | None:
    """Auto-create or update a deal from an email.

    This function creates deals based on the configured settings:
    - "auto": Create deal for every email from non-personal domains
    - "ai": Use AI to determine if email indicates a deal opportunity
    - "criteria": Only create deals for emails matching specific criteria

    Returns the created/updated deal record, or None if skipped.
    """
    from uuid import uuid4

    if not deal_settings.get("auto_create_deals", False):
        return None

    from_email = synced_email.from_email
    if not from_email:
        return None

    # Extract domain
    domain = from_email.split("@")[1].lower() if "@" in from_email else None
    if not domain:
        return None

    # Skip personal domains if configured
    personal_domains = {"gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "icloud.com", "aol.com", "protonmail.com"}
    if deal_settings.get("skip_personal_domains", True) and domain in personal_domains:
        return None

    # Get creation mode
    mode = deal_settings.get("deal_creation_mode", "auto")

    # For criteria mode, check if email matches
    if mode == "criteria":
        criteria = deal_settings.get("criteria", {})
        subject_keywords = [kw.lower() for kw in criteria.get("subject_keywords", [])]
        body_keywords = [kw.lower() for kw in criteria.get("body_keywords", [])]
        from_domains = [d.lower() for d in criteria.get("from_domains", [])]

        matches = False

        # Check domain filter
        if from_domains and domain in from_domains:
            matches = True

        # Check subject keywords
        subject = (synced_email.subject or "").lower()
        if subject_keywords and any(kw in subject for kw in subject_keywords):
            matches = True

        # Check body keywords
        body = (synced_email.body_text or "").lower()
        if body_keywords and any(kw in body for kw in body_keywords):
            matches = True

        if not matches:
            return None

    # For AI mode, use LLM to determine if this is a deal opportunity
    if mode == "ai":
        is_deal_opportunity = await _check_deal_opportunity_with_ai(
            synced_email.subject,
            synced_email.snippet or synced_email.body_text[:500] if synced_email.body_text else None,
        )
        if not is_deal_opportunity:
            return None

    # Get the Deal object type
    deal_obj_result = await db.execute(
        select(CRMObject).where(
            CRMObject.workspace_id == workspace_id,
            CRMObject.slug == "deal",
        )
    )
    deal_obj = deal_obj_result.scalar_one_or_none()

    if not deal_obj:
        logger.debug("No deal object type configured, skipping deal creation")
        return None

    # Check for existing deal linked to this company or contact
    existing_deal = None
    if company_record or contact_record:
        # Look for deals related to the company or contact
        deal_records_result = await db.execute(
            select(CRMRecord).where(
                CRMRecord.workspace_id == workspace_id,
                CRMRecord.object_id == deal_obj.id,
            )
        )
        deal_records = deal_records_result.scalars().all()

        # Check relations to find linked deals
        for deal in deal_records:
            relations_result = await db.execute(
                select(CRMRecordRelation).where(
                    CRMRecordRelation.source_record_id == deal.id
                )
            )
            relations = relations_result.scalars().all()

            for rel in relations:
                target_id = rel.target_record_id
                if company_record and target_id == company_record.id:
                    existing_deal = deal
                    break
                if contact_record and target_id == contact_record.id:
                    existing_deal = deal
                    break
            if existing_deal:
                break

    if existing_deal:
        # Update existing deal - add email info to notes/activity
        existing_values = existing_deal.values or {}

        # Update last_activity date
        existing_values["last_activity"] = datetime.now(timezone.utc).isoformat()

        # Append email to notes if there's a notes field
        if "notes" in existing_values:
            email_note = f"\n---\nEmail from {synced_email.from_email} on {synced_email.gmail_date}: {synced_email.subject}"
            existing_values["notes"] = existing_values["notes"] + email_note
        else:
            existing_values["email_count"] = existing_values.get("email_count", 0) + 1

        existing_deal.values = existing_values
        await db.flush()
        logger.info(f"Updated existing deal {existing_deal.display_name} with new email activity")

        # Link email to deal
        await _link_email_to_record(db, synced_email, existing_deal, "deal_activity")

        return existing_deal

    # Create a new deal
    deal_name = f"Deal - {synced_email.from_name or domain}"
    if synced_email.subject:
        # Use subject for more context
        deal_name = f"{synced_email.from_name or domain}: {synced_email.subject[:50]}"

    deal_values = {
        "name": deal_name,
        "stage": deal_settings.get("default_deal_stage", "new"),
        "source": "email_sync",
        "source_email": from_email,
        "created_from_email": synced_email.id,
    }

    if deal_settings.get("default_deal_value"):
        deal_values["value"] = deal_settings["default_deal_value"]

    # Add company/contact info
    if company_record:
        deal_values["company"] = company_record.display_name
        deal_values["company_domain"] = domain

    deal_record = CRMRecord(
        id=str(uuid4()),
        workspace_id=workspace_id,
        object_id=deal_obj.id,
        display_name=deal_name[:500],
        values=deal_values,
        source="email_sync",
    )
    db.add(deal_record)
    # Update object record count
    deal_obj.record_count = (deal_obj.record_count or 0) + 1
    await db.flush()

    # Create relations to company and contact
    if company_record:
        relation = CRMRecordRelation(
            id=str(uuid4()),
            source_record_id=deal_record.id,
            target_record_id=company_record.id,
            relation_type="deal_company",
        )
        db.add(relation)

    if contact_record:
        relation = CRMRecordRelation(
            id=str(uuid4()),
            source_record_id=deal_record.id,
            target_record_id=contact_record.id,
            relation_type="deal_contact",
        )
        db.add(relation)

    # Link email to deal
    await _link_email_to_record(db, synced_email, deal_record, "deal_source")

    logger.info(f"Auto-created deal '{deal_name}' from email sync")

    return deal_record


async def _check_deal_opportunity_with_ai(
    subject: str | None,
    body_preview: str | None,
) -> bool:
    """Use AI to check if an email indicates a deal opportunity."""
    if not subject and not body_preview:
        return False

    try:
        from aexy.agents.llm import llm_completion

        prompt = f"""Analyze this email and determine if it indicates a potential business deal or sales opportunity.

Subject: {subject or 'N/A'}
Preview: {body_preview or 'N/A'}

Respond with ONLY "yes" or "no".
- "yes" if the email suggests: pricing inquiry, proposal request, product interest, partnership opportunity, quote request, demo request, or similar sales indicators
- "no" for: newsletters, notifications, support tickets, internal emails, spam, or general inquiries not related to potential deals"""

        response = await llm_completion(
            messages=[{"role": "user", "content": prompt}],
            model="haiku",  # Use fast model for quick classification
            max_tokens=10,
        )

        answer = response.strip().lower()
        return answer == "yes"

    except Exception as e:
        logger.warning(f"AI deal check failed, defaulting to create deal: {e}")
        return True  # Default to creating deal if AI fails


async def _link_email_to_record(
    db: AsyncSession,
    synced_email: "SyncedEmail",
    record: CRMRecord,
    link_type: str,
) -> None:
    """Link an email to a CRM record if not already linked."""
    existing_link_result = await db.execute(
        select(SyncedEmailRecordLink).where(
            SyncedEmailRecordLink.email_id == synced_email.id,
            SyncedEmailRecordLink.record_id == record.id,
        )
    )
    if existing_link_result.scalar_one_or_none():
        return

    from uuid import uuid4
    link = SyncedEmailRecordLink(
        id=str(uuid4()),
        email_id=synced_email.id,
        record_id=record.id,
        link_type=link_type,
        confidence=1.0,
    )
    db.add(link)


async def _find_company_for_email(
    db: AsyncSession,
    workspace_id: str,
    domain: str,
) -> CRMRecord | None:
    """Find an existing company record for a given domain."""
    company_obj_result = await db.execute(
        select(CRMObject).where(
            CRMObject.workspace_id == workspace_id,
            CRMObject.slug == "company",
        )
    )
    company_obj = company_obj_result.scalar_one_or_none()

    if not company_obj:
        return None

    company_records_result = await db.execute(
        select(CRMRecord).where(
            CRMRecord.workspace_id == workspace_id,
            CRMRecord.object_id == company_obj.id,
        )
    )
    company_records = company_records_result.scalars().all()

    for record in company_records:
        if record.values:
            record_domain = record.values.get("domain", "").lower()
            record_website = record.values.get("website", "").lower()
            if domain == record_domain or domain in record_website:
                return record

    return None


settings = get_settings()

# Gmail API URLs
GMAIL_API_BASE = "https://gmail.googleapis.com/gmail/v1"
GOOGLE_TOKEN_URL = "https://oauth2.googleapis.com/token"

# Gmail API scopes required
GMAIL_SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/gmail.send",
    "https://www.googleapis.com/auth/gmail.modify",
]


class GmailSyncError(Exception):
    """Gmail sync error."""

    pass


class GmailAuthError(GmailSyncError):
    """Gmail authentication error."""

    pass


class GmailSyncService:
    """Service for syncing Gmail emails."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def _refresh_token_if_needed(
        self, integration: GoogleIntegration
    ) -> str:
        """Refresh the access token if it's about to expire.

        Delegates to the shared oauth_token_service so refresh-token
        rotation and invalid_grant handling are centralised.
        """
        from aexy.services.oauth_token_service import (
            RefreshTokenRevokedError,
            TokenRefreshError,
            ensure_valid_google_integration_token,
        )

        if not integration.refresh_token:
            raise GmailAuthError("No refresh token available")
        try:
            return await ensure_valid_google_integration_token(self.db, integration)
        except RefreshTokenRevokedError as e:
            raise GmailAuthError(
                "Google refresh token revoked — workspace needs to reconnect"
            ) from e
        except TokenRefreshError as e:
            raise GmailAuthError("Failed to refresh Google token") from e

    async def _make_gmail_request(
        self,
        integration: GoogleIntegration,
        method: str,
        endpoint: str,
        **kwargs,
    ) -> dict:
        """Make an authenticated request to the Gmail API."""
        access_token = await self._refresh_token_if_needed(integration)

        async with httpx.AsyncClient() as client:
            response = await client.request(
                method,
                f"{GMAIL_API_BASE}{endpoint}",
                headers={"Authorization": f"Bearer {access_token}"},
                **kwargs,
            )

            if response.status_code == 401:
                raise GmailAuthError("Gmail authentication failed")

            if response.status_code >= 400:
                logger.error(f"Gmail API error: {response.status_code} - {response.text}")
                raise GmailSyncError(f"Gmail API error: {response.status_code}")

            return response.json()

    async def get_or_create_sync_cursor(
        self, integration: GoogleIntegration
    ) -> EmailSyncCursor:
        """Get or create the sync cursor for an integration."""
        result = await self.db.execute(
            select(EmailSyncCursor).where(
                EmailSyncCursor.integration_id == integration.id
            )
        )
        cursor = result.scalar_one_or_none()

        if not cursor:
            cursor = EmailSyncCursor(integration_id=integration.id)
            self.db.add(cursor)
            await self.db.flush()

        return cursor

    async def start_full_sync(
        self,
        integration: GoogleIntegration,
        max_messages: int = 500,
    ) -> dict:
        """Start a full sync of emails.

        Returns sync statistics.
        """
        if not integration.gmail_sync_enabled:
            return {"error": "Gmail sync not enabled"}

        cursor = await self.get_or_create_sync_cursor(integration)

        if cursor.full_sync_completed:
            return {"message": "Full sync already completed", "messages_synced": cursor.messages_synced}

        cursor.full_sync_started_at = datetime.now(timezone.utc)
        messages_synced = 0
        page_token = cursor.next_page_token

        try:
            # Get list of messages
            params: dict[str, Any] = {
                "maxResults": min(100, max_messages),
                "labelIds": ["INBOX"],  # Start with inbox only
            }
            if page_token:
                params["pageToken"] = page_token

            while messages_synced < max_messages:
                response = await self._make_gmail_request(
                    integration,
                    "GET",
                    "/users/me/messages",
                    params=params,
                )

                messages = response.get("messages", [])
                if not messages:
                    break

                # Fetch and store each message
                for msg_info in messages:
                    try:
                        await self._sync_message(integration, msg_info["id"])
                        messages_synced += 1
                    except Exception as e:
                        logger.error(f"Failed to sync message {msg_info['id']}: {e}")
                        continue

                    if messages_synced >= max_messages:
                        break

                # Get next page token
                page_token = response.get("nextPageToken")
                if not page_token:
                    cursor.full_sync_completed = True
                    cursor.full_sync_completed_at = datetime.now(timezone.utc)
                    break

                cursor.next_page_token = page_token
                params["pageToken"] = page_token

            # Get history ID for incremental sync
            profile_response = await self._make_gmail_request(
                integration, "GET", "/users/me/profile"
            )
            cursor.history_id = profile_response.get("historyId")
            cursor.messages_synced = (cursor.messages_synced or 0) + messages_synced
            cursor.last_sync_at = datetime.now(timezone.utc)
            cursor.last_error = None
            cursor.error_count = 0

            await self.db.flush()

            return {
                "messages_synced": messages_synced,
                "full_sync_completed": cursor.full_sync_completed,
                "history_id": cursor.history_id,
            }

        except Exception as e:
            cursor.last_error = str(e)
            cursor.error_count = (cursor.error_count or 0) + 1
            await self.db.flush()
            raise

    async def start_incremental_sync(
        self, integration: GoogleIntegration
    ) -> dict:
        """Sync new emails since the last sync using Gmail History API.

        Returns sync statistics.
        """
        if not integration.gmail_sync_enabled:
            return {"error": "Gmail sync not enabled"}

        cursor = await self.get_or_create_sync_cursor(integration)

        if not cursor.history_id:
            # No history ID - need full sync first
            return await self.start_full_sync(integration)

        messages_synced = 0
        try:
            # Get history since last sync
            response = await self._make_gmail_request(
                integration,
                "GET",
                "/users/me/history",
                params={
                    "startHistoryId": cursor.history_id,
                    "historyTypes": ["messageAdded"],
                },
            )

            history = response.get("history", [])
            new_history_id = response.get("historyId")

            # Process each history record
            seen_message_ids = set()
            for record in history:
                for msg_added in record.get("messagesAdded", []):
                    msg_id = msg_added["message"]["id"]
                    if msg_id not in seen_message_ids:
                        seen_message_ids.add(msg_id)
                        try:
                            await self._sync_message(integration, msg_id)
                            messages_synced += 1
                        except Exception as e:
                            logger.error(f"Failed to sync message {msg_id}: {e}")

            # Update cursor
            if new_history_id:
                cursor.history_id = new_history_id
            cursor.last_sync_at = datetime.now(timezone.utc)
            cursor.last_error = None

            await self.db.flush()

            return {
                "messages_synced": messages_synced,
                "history_id": cursor.history_id,
            }

        except GmailSyncError as e:
            if "historyId" in str(e).lower():
                # History ID expired - need full sync
                cursor.history_id = None
                cursor.full_sync_completed = False
                await self.db.flush()
                return await self.start_full_sync(integration)
            raise

    async def _sync_message(
        self, integration: GoogleIntegration, message_id: str
    ) -> SyncedEmail:
        """Fetch and store a single message."""
        # Check if already synced
        result = await self.db.execute(
            select(SyncedEmail).where(SyncedEmail.gmail_id == message_id)
        )
        existing = result.scalar_one_or_none()
        if existing:
            return existing

        # Fetch full message
        message = await self._make_gmail_request(
            integration,
            "GET",
            f"/users/me/messages/{message_id}",
            params={"format": "full"},
        )

        # Parse message
        email_data = self._parse_message(message)

        # Create synced email record
        synced_email = SyncedEmail(
            workspace_id=integration.workspace_id,
            integration_id=integration.id,
            gmail_id=message_id,
            gmail_thread_id=message.get("threadId"),
            subject=email_data.get("subject"),
            from_email=email_data.get("from_email"),
            from_name=email_data.get("from_name"),
            to_emails=email_data.get("to_emails"),
            cc_emails=email_data.get("cc_emails"),
            snippet=message.get("snippet"),
            body_text=email_data.get("body_text"),
            body_html=email_data.get("body_html"),
            labels=message.get("labelIds"),
            is_read="UNREAD" not in (message.get("labelIds") or []),
            is_starred="STARRED" in (message.get("labelIds") or []),
            has_attachments=email_data.get("has_attachments", False),
            gmail_date=email_data.get("date"),
        )

        self.db.add(synced_email)
        await self.db.flush()

        # Service Desk intake: if this mailbox is a registered service-desk
        # source, turn the email into a ticket and skip CRM enrichment (a
        # partner writing the ops desk is not a sales contact).
        try:
            from aexy.services.service_desk_intake_service import (
                ServiceDeskIntakeService,
                ai_classification_enabled,
            )
            from aexy.schemas.service_desk import InboundEmail

            mailbox = await ServiceDeskIntakeService.find_mailbox_by_integration(
                self.db, integration.id
            )
            if mailbox is not None:
                try:
                    attachments = await self._service_desk_attachment_context(
                        integration,
                        message_id,
                        message.get("payload") or {},
                        with_previews=await ai_classification_enabled(
                            self.db, mailbox.workspace_id
                        ),
                    )
                except Exception as exc:  # noqa: BLE001 - context must not block intake
                    logger.info("Service desk attachment context skipped: %s", exc)
                    attachments = []
                to_addr = None
                if email_data.get("to_emails"):
                    first = email_data["to_emails"][0]
                    to_addr = first.get("email") if isinstance(first, dict) else first
                intake = ServiceDeskIntakeService(self.db)
                # Savepoint: a failed intake must not poison the session for the
                # rest of this sync. Swallowing the error and carrying on with a
                # transaction that needs rollback made every later statement in
                # the batch fail with PendingRollbackError.
                async with self.db.begin_nested():
                    await intake.ingest(
                        InboundEmail(
                            to=to_addr or mailbox.address,
                            from_email=email_data.get("from_email") or "",
                            from_name=email_data.get("from_name"),
                            subject=email_data.get("subject") or "",
                            body_text=email_data.get("body_text") or "",
                            body_html=email_data.get("body_html"),
                            message_id=message_id,
                            thread_id=message.get("threadId"),
                            attachments=attachments,
                            headers=email_data.get("headers") or {},
                        ),
                        mailbox,
                        source="service_desk_gmail",
                    )
                # Each message is an independent unit; commit before notifying so
                # the requester is never acknowledged for a rolled-back ticket.
                await self.db.commit()
                await intake.flush_notifications()
                return synced_email
        except Exception as e:
            logger.warning(f"Service desk intake (gmail) failed: {e}")

        # Auto-enrich: create contact and company from email sender if not exists
        contact_record = None
        company_record = None
        try:
            contact_record, company_record = await auto_enrich_contact_from_email(
                db=self.db,
                workspace_id=integration.workspace_id,
                from_email=email_data.get("from_email"),
                from_name=email_data.get("from_name"),
                synced_email=synced_email,
            )
        except Exception as e:
            # Don't fail sync if enrichment fails
            logger.warning(f"Failed to auto-enrich contact from email: {e}")

        # Auto-create deal from email if enabled
        try:
            sync_settings = integration.sync_settings or {}
            deal_settings = {**DEFAULT_DEAL_SETTINGS, **sync_settings.get("deal_settings", {})}

            if deal_settings.get("auto_create_deals", False):
                await auto_create_or_update_deal_from_email(
                    db=self.db,
                    workspace_id=integration.workspace_id,
                    synced_email=synced_email,
                    contact_record=contact_record,
                    company_record=company_record,
                    deal_settings=deal_settings,
                )
        except Exception as e:
            # Don't fail sync if deal creation fails
            logger.warning(f"Failed to auto-create deal from email: {e}")

        return synced_email

    def _parse_message(self, message: dict) -> dict:
        """Parse Gmail message into structured data."""
        payload = message.get("payload", {})
        headers = {h["name"].lower(): h["value"] for h in payload.get("headers", [])}

        # Parse from address
        from_header = headers.get("from", "")
        from_name, from_email = parseaddr(from_header)

        # Parse to addresses
        to_header = headers.get("to", "")
        to_emails = self._parse_email_list(to_header)

        # Parse CC addresses
        cc_header = headers.get("cc", "")
        cc_emails = self._parse_email_list(cc_header)

        # Parse date
        date_str = headers.get("date", "")
        date = None
        if date_str:
            try:
                from email.utils import parsedate_to_datetime
                date = parsedate_to_datetime(date_str)
            except Exception:
                pass

        # Get body
        body_text, body_html = self._extract_body(payload)

        # Check for attachments
        has_attachments = self._has_attachments(payload)

        return {
            "subject": headers.get("subject"),
            "from_email": from_email,
            "from_name": from_name,
            "to_emails": to_emails,
            "cc_emails": cc_emails,
            "date": date,
            "body_text": body_text,
            "body_html": body_html,
            "has_attachments": has_attachments,
            # Service Desk intake reads these to recognise auto-responders and
            # the mail this application itself sent from this same account.
            "headers": headers,
        }

    def _parse_email_list(self, header_value: str) -> list[dict]:
        """Parse a comma-separated email list into structured data."""
        if not header_value:
            return []

        emails = []
        # Split by comma but handle quoted names
        parts = header_value.split(",")
        for part in parts:
            name, email = parseaddr(part.strip())
            if email:
                emails.append({"name": name or None, "email": email})
        return emails

    def _extract_body(self, payload: dict) -> tuple[str | None, str | None]:
        """Extract text and HTML body from message payload."""
        body_text = None
        body_html = None

        def process_part(part: dict):
            nonlocal body_text, body_html
            mime_type = part.get("mimeType", "")
            body = part.get("body", {})
            data = body.get("data")

            # A part carrying a filename is an attachment, not message body: a
            # text/plain or text/html attachment used to be base64-decoded whole
            # right here, bypassing the raw-byte ceiling that guards the Service
            # Desk attachment path — and then landing in body_text as if the
            # requester had typed it.
            if data and not part.get("filename"):
                decoded = base64.urlsafe_b64decode(data).decode("utf-8", errors="ignore")
                if mime_type == "text/plain" and not body_text:
                    body_text = decoded
                elif mime_type == "text/html" and not body_html:
                    body_html = decoded

            # Process nested parts
            for sub_part in part.get("parts", []):
                process_part(sub_part)

        process_part(payload)
        return body_text, body_html

    def _has_attachments(self, payload: dict) -> bool:
        """Check if message has attachments."""
        def check_part(part: dict) -> bool:
            if part.get("filename"):
                return True
            for sub_part in part.get("parts", []):
                if check_part(sub_part):
                    return True
            return False

        return check_part(payload)

    async def _service_desk_attachment_context(
        self,
        integration: GoogleIntegration,
        message_id: str,
        payload: dict,
        with_previews: bool = True,
    ) -> list[dict[str, Any]]:
        """Return attachment metadata plus small, best-effort classifier previews.

        ``with_previews`` is the workspace's AI opt-in. The previews exist only
        to feed the classifier, so with AI off none are built and no attachment
        is ever downloaded — the KAM still sees the file name, type and size.
        """
        parts: list[dict] = []

        def collect(part: dict) -> None:
            if part.get("filename"):
                parts.append(part)
            for nested in part.get("parts", []):
                if isinstance(nested, dict):
                    collect(nested)

        collect(payload)
        context: list[dict[str, Any]] = []
        previews_loaded = 0
        for part in parts[:_SERVICE_DESK_ATTACHMENT_METADATA_LIMIT]:
            body = part.get("body") if isinstance(part.get("body"), dict) else {}
            filename = str(part.get("filename") or "attachment")
            content_type = str(part.get("mimeType") or "application/octet-stream")
            size_bytes = self._declared_attachment_size(body)
            item: dict[str, Any] = {
                "filename": filename,
                "content_type": content_type,
                "size_bytes": size_bytes,
                # Identifier only, never content, so it is captured whether or not
                # AI is on. Without it the desk can show that a claim register
                # arrived but can never forward it to the insurer.
                "attachment_id": (body.get("attachmentId") if isinstance(body, dict) else None),
            }
            if (
                with_previews
                and previews_loaded < _SERVICE_DESK_ATTACHMENT_PREVIEW_LIMIT
                and self._supports_service_desk_preview(filename, content_type)
            ):
                try:
                    raw = await self._gmail_attachment_bytes(
                        integration,
                        message_id,
                        body,
                    )
                    item["preview"] = self._service_desk_preview(
                        filename,
                        content_type,
                        raw,
                    )
                    if item["preview"]:
                        previews_loaded += 1
                    if item["size_bytes"] is None:
                        item["size_bytes"] = len(raw)
                except Exception as exc:  # noqa: BLE001 - attachment context is best-effort
                    logger.info(
                        "Service desk attachment preview skipped for %s: %s",
                        filename,
                        exc,
                    )
            context.append(item)
        return context

    async def _gmail_attachment_bytes(
        self,
        integration: GoogleIntegration,
        message_id: str,
        body: dict,
        max_bytes: int | None = None,
    ) -> bytes:
        """Load attachment bytes only when every pre-decode size check is safe.

        ``max_bytes`` defaults to the classifier's preview ceiling. Forwarding
        passes its own, larger one: a claim register too big to sample is still
        a file the insurer needs to receive.
        """
        limit = max_bytes or _SERVICE_DESK_ATTACHMENT_RAW_BYTE_LIMIT
        declared_size = self._declared_attachment_size(body)
        if declared_size is not None and declared_size > limit:
            raise ValueError("attachment exceeds the Service Desk raw-byte limit")

        encoded = body.get("data")
        if not encoded and body.get("attachmentId"):
            if declared_size is None:
                raise ValueError("attachment size is unavailable before download")
            response = await self._make_gmail_request(
                integration,
                "GET",
                f"/users/me/messages/{message_id}/attachments/{body['attachmentId']}",
            )
            response_size = self._declared_attachment_size(response)
            if response_size is not None and response_size > limit:
                raise ValueError("attachment response exceeds the Service Desk raw-byte limit")
            encoded = response.get("data")
        if not encoded:
            return b""

        if isinstance(encoded, bytes):
            encoded_text = encoded.decode("ascii")
        else:
            encoded_text = str(encoded)
        encoded_without_padding = encoded_text.rstrip("=")
        decoded_size_upper_bound = (len(encoded_without_padding) * 3) // 4
        if decoded_size_upper_bound > limit:
            raise ValueError("encoded attachment exceeds the Service Desk raw-byte limit")

        raw = base64.urlsafe_b64decode(
            encoded_text + "=" * (-len(encoded_text) % 4)
        )
        if len(raw) > _SERVICE_DESK_ATTACHMENT_RAW_BYTE_LIMIT:
            raise ValueError("decoded attachment exceeds the Service Desk raw-byte limit")
        return self._decode_if_still_base64(raw)

    @staticmethod
    def _decode_if_still_base64(raw: bytes) -> bytes:
        """Undo a second layer of base64 left by some senders' MIME parts.

        Gmail normally hands back the part's decoded content, but for some
        clients the transfer encoding survives and what arrives is the base64
        text itself. Everything downstream then breaks in the same way: the
        classifier reads an unintelligible blob instead of the claim rows, and
        the file the desk forwards to an insurer cannot be opened. Decoding here
        fixes both, because every caller loads its bytes through this method.

        Only applied when the payload is nothing but base64 characters, so real
        files are untouched: any CSV has commas, any prose has spaces, and any
        binary format has bytes outside the alphabet.
        """
        compact = raw.translate(None, b"\r\n")
        if len(compact) < 8 or len(compact) % 4 != 0:
            return raw
        if not _BASE64_ONLY_RE.fullmatch(compact):
            return raw
        try:
            return base64.b64decode(compact, validate=True)
        except Exception:  # noqa: BLE001 — an undecodable payload is just not double-encoded
            return raw

    @staticmethod
    def _declared_attachment_size(body: dict) -> int | None:
        value = body.get("size")
        try:
            size = int(value)
        except (TypeError, ValueError):
            return None
        return size if size >= 0 else None

    @staticmethod
    def _supports_service_desk_preview(filename: str, content_type: str) -> bool:
        name = filename.lower()
        return (
            name.endswith((".csv", ".tsv", ".xlsx", ".txt", ".md", ".log", ".json", ".xml"))
            or content_type.startswith("text/")
            or content_type in {"application/json", "application/xml"}
        )

    @staticmethod
    def _xlsx_expansion_is_bounded(raw: bytes) -> bool:
        try:
            with zipfile.ZipFile(io.BytesIO(raw)) as archive:
                expanded_size = 0
                for member in archive.infolist():
                    expanded_size += member.file_size
                    if expanded_size > _SERVICE_DESK_XLSX_EXPANDED_BYTE_LIMIT:
                        return False
        except zipfile.BadZipFile:
            return False
        return True

    @staticmethod
    def _service_desk_preview(filename: str, content_type: str, raw: bytes) -> str | None:
        """Produce a capped spreadsheet sample or a short text preview."""
        if not raw or len(raw) > _SERVICE_DESK_ATTACHMENT_RAW_BYTE_LIMIT:
            return None
        name = filename.lower()
        try:
            if name.endswith(".xlsx"):
                if not GmailSyncService._xlsx_expansion_is_bounded(raw):
                    return None
                workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                try:
                    rows = list(workbook.active.iter_rows(values_only=True, max_row=4))
                finally:
                    workbook.close()
                return json.dumps(
                    [
                        [str(cell)[:120] if cell is not None else "" for cell in row]
                        for row in rows
                    ],
                    ensure_ascii=False,
                )[:_SERVICE_DESK_ATTACHMENT_PREVIEW_CHAR_LIMIT]
            if name.endswith((".csv", ".tsv")):
                delimiter = "\t" if name.endswith(".tsv") else ","
                reader = csv.reader(
                    io.StringIO(raw.decode("utf-8", errors="replace")),
                    delimiter=delimiter,
                )
                rows = list(islice(reader, 4))
                return json.dumps(
                    [[cell[:120] for cell in row] for row in rows],
                    ensure_ascii=False,
                )[:_SERVICE_DESK_ATTACHMENT_PREVIEW_CHAR_LIMIT]
            text = raw.decode("utf-8", errors="replace")
            return (
                " ".join(text.split())[:_SERVICE_DESK_ATTACHMENT_PREVIEW_CHAR_LIMIT]
                or None
            )
        except Exception as exc:  # noqa: BLE001 - parsing must never block ticket intake
            logger.info(
                "Service desk attachment preview parsing skipped for %s: %s",
                filename,
                exc,
            )
            return None

    async def link_emails_to_records(
        self,
        workspace_id: str,
        email_ids: list[str] | None = None,
    ) -> dict:
        """Link synced emails to CRM records by email address matching."""
        # Get emails to process
        query = select(SyncedEmail).where(SyncedEmail.workspace_id == workspace_id)
        if email_ids:
            query = query.where(SyncedEmail.id.in_(email_ids))

        result = await self.db.execute(query)
        emails = result.scalars().all()

        # Get all records with email attributes
        records_result = await self.db.execute(
            select(CRMRecord)
            .where(CRMRecord.workspace_id == workspace_id)
            .options(selectinload(CRMRecord.object))
        )
        records = records_result.scalars().all()

        # Build email -> record mapping
        email_to_records: dict[str, list[CRMRecord]] = {}
        for record in records:
            if not record.values:
                continue
            # Look for email fields in record values
            for key, value in record.values.items():
                if isinstance(value, str) and "@" in value:
                    email_lower = value.lower()
                    if email_lower not in email_to_records:
                        email_to_records[email_lower] = []
                    email_to_records[email_lower].append(record)

        links_created = 0
        for email in emails:
            # Check from email
            if email.from_email:
                for record in email_to_records.get(email.from_email.lower(), []):
                    await self._create_email_link(email, record, "from")
                    links_created += 1

            # Check to emails
            for to in email.to_emails or []:
                to_email = to.get("email", "").lower()
                for record in email_to_records.get(to_email, []):
                    await self._create_email_link(email, record, "to")
                    links_created += 1

        await self.db.flush()
        return {"links_created": links_created}

    async def _create_email_link(
        self,
        email: SyncedEmail,
        record: CRMRecord,
        link_type: str,
    ) -> SyncedEmailRecordLink | None:
        """Create a link between an email and a record if it doesn't exist."""
        # Check if link already exists
        result = await self.db.execute(
            select(SyncedEmailRecordLink).where(
                SyncedEmailRecordLink.email_id == email.id,
                SyncedEmailRecordLink.record_id == record.id,
            )
        )
        if result.scalar_one_or_none():
            return None

        link = SyncedEmailRecordLink(
            email_id=email.id,
            record_id=record.id,
            link_type=link_type,
            confidence=1.0,
        )
        self.db.add(link)
        return link

    async def send_email(
        self,
        integration: GoogleIntegration,
        to: str,
        subject: str,
        body_html: str,
        reply_to_message_id: str | None = None,
    ) -> dict:
        """Send an email via Gmail API.

        Returns the sent message info.
        """
        import email.mime.text
        import email.mime.multipart

        # Build the email
        msg = email.mime.multipart.MIMEMultipart("alternative")
        msg["To"] = to
        msg["Subject"] = subject
        msg["From"] = integration.google_email

        if reply_to_message_id:
            msg["In-Reply-To"] = reply_to_message_id
            msg["References"] = reply_to_message_id

        # Add HTML body
        html_part = email.mime.text.MIMEText(body_html, "html")
        msg.attach(html_part)

        # Encode message
        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")

        # Send via Gmail API
        response = await self._make_gmail_request(
            integration,
            "POST",
            "/users/me/messages/send",
            json={"raw": raw},
        )

        return {
            "message_id": response.get("id"),
            "thread_id": response.get("threadId"),
        }
